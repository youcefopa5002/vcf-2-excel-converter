# -*- coding: utf-8 -*-
"""
================================================================================
## vCard to Excel Converter (GUI Version) ##

A modern and stylish desktop application to convert vCard files (.vcf) into 
structured Excel files (.xlsx).

--------------------------------------------------------------------------------
### Features:
- Clean, modern, dark-themed graphical interface.
- Simple one-click file selection.
- "Save As" dialog to choose the output location and name.
- Fully automatic and robust file processing, handling corrupted files
  and various encodings without user input.
- Real-time status updates.

--------------------------------------------------------------------------------
### How to Run:
1. Make sure you have the required libraries installed:
   `pip install vobject openpyxl phonenumbers chardet`
2. Run the script:
   `python vcf_converter_gui.py`
================================================================================
"""

import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, font
from typing import List, Set, Tuple, Optional

# --- Backend Libraries ---
import chardet
import openpyxl
import phonenumbers
import vobject
from openpyxl.styles import Font as OpenPyxlFont
from openpyxl.utils import get_column_letter


# --- Backend Logic (Core conversion functions) ---

def read_vcf_content(file_path: str) -> str:
    """
    Reads VCF file content with maximum robustness by reading as binary first
    and ignoring any decoding errors. This is the final fallback for corrupted files.
    """
    try:
        with open(file_path, 'rb') as f:
            raw_data = f.read()
    except (IOError, FileNotFoundError) as e:
        raise ValueError(f"Could not read the file from disk: {e}")

    # --- Automatic detection logic ---
    encodings_to_try = ['utf-8', 'cp1256']
    
    try:
        detected = chardet.detect(raw_data)['encoding']
        if detected and detected.lower() not in encodings_to_try:
            encodings_to_try.append(detected.lower())
    except Exception:
        pass

    encodings_to_try.extend(['utf-8-sig', 'latin-1', 'iso-8859-6', 'cp1252'])

    for encoding in encodings_to_try:
        try:
            return raw_data.decode(encoding, errors='ignore')
        except (UnicodeError, TypeError):
            continue
            
    return raw_data.decode('utf-8', errors='replace')


def normalize_phone_number(number_str: str, region: str) -> str:
    """Normalizes a phone number to E.164 format."""
    if not number_str:
        return ""
    try:
        cleaned_number = "".join(filter(lambda char: char.isdigit() or char in '+', number_str))
        parsed_number = phonenumbers.parse(cleaned_number, region)
        if phonenumbers.is_valid_number(parsed_number):
            return phonenumbers.format_number(parsed_number, phonenumbers.PhoneNumberFormat.E164)
    except phonenumbers.phonenumbersutil.NumberParseException:
        pass
    return number_str.strip()

def get_contact_name(vcard: vobject.vCard) -> str:
    """Extracts the contact name using a fallback strategy: FN -> N -> ORG."""
    if hasattr(vcard, 'fn'):
        if vcard.fn.value and vcard.fn.value.strip():
            return vcard.fn.value.strip()
    if hasattr(vcard, 'n'):
        n = vcard.n.value
        name_parts = [part.strip() for part in [n.given, n.middle, n.family] if part and part.strip()]
        if name_parts:
            return " ".join(name_parts)
    if hasattr(vcard, 'org'):
        if vcard.org.value and vcard.org.value[0] and vcard.org.value[0].strip():
            return vcard.org.value[0].strip()
    return "غير معروف"

def process_vcf_data(vcf_content: str, default_region: str) -> List[Tuple[str, str]]:
    """Parses VCF content and extracts a list of (name, phone_number) tuples."""
    processed_contacts = []
    try:
        for card_text in vcf_content.split('BEGIN:VCARD'):
            if 'END:VCARD' in card_text:
                full_card = 'BEGIN:VCARD\n' + card_text
                try:
                    vcard = vobject.readOne(full_card)
                    name = get_contact_name(vcard)
                    normalized_numbers: Set[str] = set()
                    if hasattr(vcard, 'tel_list'):
                        for tel in vcard.tel_list:
                            if tel.value and tel.value.strip():
                                normalized = normalize_phone_number(tel.value, default_region)
                                if normalized:
                                    normalized_numbers.add(normalized)
                    
                    if not normalized_numbers:
                        continue
                        
                    for number in sorted(list(normalized_numbers)):
                        processed_contacts.append((name, number))
                except Exception:
                    continue

    except Exception as e:
        raise ValueError(f"Failed to parse VCF data. The file might be corrupted. Details: {e}")

    return processed_contacts

def create_excel_file(data: List[Tuple[str, str]], output_path: str):
    """Creates an XLSX file from the processed contact data."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Contacts"
    sheet.sheet_view.rightToLeft = True

    header = ["الاسم", "الرقم"]
    sheet.append(header)
    
    header_font = OpenPyxlFont(bold=True)
    for cell in sheet[1]:
        cell.font = header_font
    
    sheet.freeze_panes = 'A2'
    
    for name, number in data:
        sheet.append([name, number])
        
    for cell in sheet['B']:
        cell.number_format = '@'

    for column_cells in sheet.columns:
        max_length = 0
        column = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    workbook.save(output_path)


# --- Frontend Logic (GUI Application Class) ---

class VcfConverterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("✨ محول vCard إلى Excel ✨")
        self.root.geometry("550x380")
        self.root.resizable(False, False)
        self.root.configure(bg="#1E1E1E")

        self.input_filepath = None
        
        # --- Modern Fonts ---
        self.font_main = font.Font(family="Segoe UI", size=12)
        self.font_bold = font.Font(family="Segoe UI Semibold", size=14)
        self.font_status = font.Font(family="Segoe UI", size=10)

        # --- Color Palette ---
        self.COLOR_BG = "#1E1E1E"
        self.COLOR_FRAME = "#2D2D2D"
        self.COLOR_TEXT = "#FFFFFF"
        self.COLOR_ACCENT = "#007ACC"
        self.COLOR_SUCCESS = "#28a745"
        self.COLOR_ERROR = "#d9534f"
        
        # --- Main container frame ---
        main_frame = tk.Frame(root, bg=self.COLOR_BG, padx=30, pady=30)
        main_frame.pack(fill="both", expand=True)

        # --- Header ---
        header_label = tk.Label(main_frame, text="تحويل جهات الاتصال VCF إلى Excel", font=self.font_bold, bg=self.COLOR_BG, fg=self.COLOR_TEXT)
        header_label.pack(pady=(0, 20))

        # --- File Selection Area ---
        file_frame = tk.Frame(main_frame, bg=self.COLOR_FRAME, relief="solid", bd=1)
        file_frame.pack(fill="x", pady=20, ipady=10, ipadx=10)

        self.select_btn = tk.Button(file_frame, text="📂  اختيار ملف VCF", command=self.select_file, 
                                     font=self.font_main, bg=self.COLOR_ACCENT, fg=self.COLOR_TEXT, 
                                     relief="flat", width=18, activebackground="#005f9e", activeforeground="white", cursor="hand2")
        self.select_btn.pack(side="left", padx=(10, 10))

        self.file_label = tk.Label(file_frame, text="...لم يتم اختيار أي ملف", bg=self.COLOR_FRAME, 
                                    fg="#a0a0a0", anchor="w", padx=10, font=self.font_main)
        self.file_label.pack(side="left", fill="x", expand=True)
        
        # --- Convert Button ---
        self.convert_btn = tk.Button(main_frame, text="🚀  بدء التحويل", command=self.convert, 
                                      font=self.font_bold, bg=self.COLOR_SUCCESS, fg=self.COLOR_TEXT, 
                                      relief="flat", height=2, state="disabled",
                                      activebackground="#218838", activeforeground="white", cursor="hand2")
        self.convert_btn.pack(fill="x", pady=(20, 10))
        
        # --- Status Bar ---
        self.status_label = tk.Label(root, text="جاهز للبدء", bd=1, relief="sunken", anchor="w", 
                                      padx=10, font=self.font_status, bg="#252526", fg=self.COLOR_TEXT)
        self.status_label.pack(side="bottom", fill="x")

        # --- Hover Effects ---
        self.select_btn.bind("<Enter>", lambda e: self.select_btn.config(bg="#005f9e"))
        self.select_btn.bind("<Leave>", lambda e: self.select_btn.config(bg=self.COLOR_ACCENT))
        
        def on_convert_enter(e):
            if self.convert_btn['state'] == 'normal': self.convert_btn.config(bg="#218838")
        def on_convert_leave(e):
            if self.convert_btn['state'] == 'normal': self.convert_btn.config(bg=self.COLOR_SUCCESS)

        self.convert_btn.bind("<Enter>", on_convert_enter)
        self.convert_btn.bind("<Leave>", on_convert_leave)

    def select_file(self):
        filepath = filedialog.askopenfilename(
            title="اختر ملف vCard",
            filetypes=[("vCard Files", "*.vcf"), ("All Files", "*.*")]
        )
        if filepath:
            self.input_filepath = filepath
            filename = os.path.basename(filepath)
            self.file_label.config(text=filename, fg=self.COLOR_TEXT)
            self.convert_btn.config(state="normal", bg=self.COLOR_SUCCESS)
            self.update_status(f"تم اختيار الملف: {filename}", "black")

    def convert(self):
        if not self.input_filepath:
            messagebox.showerror("خطأ", "الرجاء اختيار ملف VCF أولاً.")
            return

        default_region = "DZ"
        
        self.update_status("...جاري المعالجة، يرجى الانتظار", "blue")
        self.root.update_idletasks()

        try:
            vcf_content = read_vcf_content(self.input_filepath)
            contact_data = process_vcf_data(vcf_content, default_region)

            if not contact_data:
                self.update_status("لم يتم العثور على أي جهات اتصال تحتوي على أرقام هواتف.", "orange")
                messagebox.showwarning("تنبيه", "لم يتم العثور على أي جهات اتصال تحتوي على أرقام هواتف في الملف المحدد.")
                return

            output_path = filedialog.asksaveasfilename(
                title="حفظ ملف Excel باسم",
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile=f"{os.path.splitext(os.path.basename(self.input_filepath))[0]}.xlsx"
            )

            if not output_path:
                self.update_status("تم إلغاء العملية.", "gray")
                return

            create_excel_file(contact_data, output_path)
            
            self.update_status(f"تم التحويل بنجاح! {len(contact_data)} رقم تم حفظه.", "green")
            messagebox.showinfo("نجاح", f"تم حفظ الملف بنجاح في:\n{output_path}")

        except Exception as e:
            self.update_status(f"خطأ: {e}", "red")
            messagebox.showerror("حدث خطأ", str(e))

    def update_status(self, message, color_name):
        color_map = {
            "black": self.COLOR_TEXT, "blue": "#5bc0de", "orange": "#f0ad4e",
            "green": self.COLOR_SUCCESS, "red": self.COLOR_ERROR, "gray": "#a0a0a0"
        }
        color = color_map.get(color_name, self.COLOR_TEXT)
        self.status_label.config(text=message, fg=color)


if __name__ == '__main__':
    root = tk.Tk()
    app = VcfConverterApp(root)
    root.mainloop()

